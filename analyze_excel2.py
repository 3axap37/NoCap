"""
Comprehensive Excel analysis - only prints non-trivial / non-default data
to keep output manageable.
"""
import openpyxl
from openpyxl.utils import get_column_letter

FILE = r"C:\W\Coding\PjtKai\CapTableExample.xlsx"

DEFAULT_FONT_NAMES = {"Calibri", "Arial", "Times New Roman", "\ub9d1\uc740 \uace0\ub515", None}
DEFAULT_RGB = {"00000000", "FF000000", "FFFFFFFF"}

def color_str(color):
    if color is None:
        return "None"
    try:
        t = color.type
        if t == "rgb":
            return f"rgb:{color.rgb}"
        elif t == "theme":
            return f"theme:{color.theme},tint:{color.tint}"
        elif t == "indexed":
            return f"indexed:{color.indexed}"
        else:
            return f"type={t}"
    except Exception as e:
        return f"(err:{e})"

def fill_str(fill):
    if fill is None:
        return None
    try:
        ft = getattr(fill, "fill_type", None) or getattr(fill, "patternType", None)
        fg_raw = fill.fgColor.rgb if hasattr(fill, "fgColor") else "00000000"
        bg_raw = fill.bgColor.rgb if hasattr(fill, "bgColor") else "00000000"
        fg = color_str(fill.fgColor) if hasattr(fill, "fgColor") else "N/A"
        bg = color_str(fill.bgColor) if hasattr(fill, "bgColor") else "N/A"
        # Skip default/empty fills
        if ft in (None, "none") and fg_raw in DEFAULT_RGB and bg_raw in DEFAULT_RGB:
            return None
        return f"fill_type={ft}, fgColor={fg}, bgColor={bg}"
    except Exception as e:
        return f"(err:{e})"

def font_str(font):
    if font is None:
        return None
    try:
        parts = []
        if font.bold:
            parts.append("BOLD")
        if font.italic:
            parts.append("italic")
        col = color_str(font.color) if font.color else None
        if col and col not in ("rgb:FF000000", "rgb:00000000", "theme:1,tint:0.0", "None"):
            parts.append(f"color={col}")
        if font.size and font.size != 11.0:
            parts.append(f"size={font.size}")
        if parts:
            return ", ".join(parts)
        return None
    except Exception as e:
        return f"(err:{e})"

def analyze_sheet_formulas(ws):
    print(f"\n{'='*70}")
    print(f"SHEET: {ws.title}")
    print(f"{'='*70}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Rows: {ws.min_row}-{ws.max_row}  Cols: {ws.min_column}-{ws.max_column}")

    print(f"\n--- Non-Empty Cells (with value/formula) ---")
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            coord = cell.coordinate
            fill = fill_str(cell.fill)
            font = font_str(cell.font)
            numfmt = cell.number_format if cell.number_format != "General" else None
            parts = [f"value={repr(v)}"]
            if fill:
                parts.append(f"fill=({fill})")
            if font:
                parts.append(f"font=({font})")
            if numfmt:
                parts.append(f"numfmt={numfmt!r}")
            print(f"  [{coord}] " + " | ".join(parts))

    print(f"\n--- ALL Cells with Notable Formatting (even if empty) ---")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                continue  # already shown above
            fill = fill_str(cell.fill)
            font = font_str(cell.font)
            numfmt = cell.number_format if cell.number_format != "General" else None
            if fill or font or numfmt:
                coord = cell.coordinate
                parts = ["(empty)"]
                if fill:
                    parts.append(f"fill=({fill})")
                if font:
                    parts.append(f"font=({font})")
                if numfmt:
                    parts.append(f"numfmt={numfmt!r}")
                print(f"  [{coord}] " + " | ".join(parts))

    print(f"\n--- Merged Cell Ranges ---")
    if ws.merged_cells.ranges:
        for rng in ws.merged_cells.ranges:
            print(f"  {rng}")
    else:
        print("  (none)")

    print(f"\n--- Column Widths (custom only) ---")
    shown = False
    for col_letter, cd in ws.column_dimensions.items():
        if cd.width or cd.customWidth or cd.hidden:
            print(f"  Col {col_letter}: width={cd.width}, customWidth={cd.customWidth}, hidden={cd.hidden}")
            shown = True
    if not shown:
        print("  (none set)")

    print(f"\n--- Row Heights (non-default only) ---")
    shown = False
    for row_idx, rd in ws.row_dimensions.items():
        if rd.height or rd.customHeight or rd.hidden:
            print(f"  Row {row_idx}: height={rd.height}, customHeight={rd.customHeight}, hidden={rd.hidden}")
            shown = True
    if not shown:
        print("  (none set)")

    print(f"\n--- Data Validation ---")
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            print(f"  type={dv.type}, formula1={dv.formula1!r}, formula2={dv.formula2!r}, "
                  f"sqref={dv.sqref}, showErrorMessage={dv.showErrorMessage}, "
                  f"errorTitle={dv.errorTitle!r}, error={dv.error!r}")
    else:
        print("  (none)")


# ============================================================
# Pass 1: data_only=False  (formulas)
# ============================================================
print("\n" + "#"*70)
print("PASS 1: data_only=False  (formulas)")
print("#"*70)

wb_f = openpyxl.load_workbook(FILE, data_only=False)
print(f"\nSheet names: {wb_f.sheetnames}")

print(f"\n--- Named Ranges (workbook level) ---")
try:
    named = list(wb_f.defined_names)
    if named:
        for dn in named:
            print(f"  name={dn.name!r}, value={dn.value!r}, localSheetId={dn.localSheetId}")
    else:
        print("  (none)")
except Exception as e:
    print(f"  (error reading named ranges: {e})")

for ws in wb_f.worksheets:
    analyze_sheet_formulas(ws)

# ============================================================
# Pass 2: data_only=True  (computed values)
# ============================================================
print("\n\n" + "#"*70)
print("PASS 2: data_only=True  (computed/cached values for formula cells)")
print("#"*70)

wb_v = openpyxl.load_workbook(FILE, data_only=True)

for ws in wb_v.worksheets:
    print(f"\n{'='*70}")
    print(f"SHEET: {ws.title}  [computed values]")
    print(f"{'='*70}")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(f"  [{cell.coordinate}] {repr(cell.value)}")

print("\n\nDONE.")
