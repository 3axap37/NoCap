import openpyxl
from openpyxl.utils import get_column_letter

FILE = r"C:\W\Coding\PjtKai\CapTableExample.xlsx"

def color_str(color):
    if color is None:
        return "None"
    try:
        t = color.type
        if t == "rgb":
            return f"rgb:{color.rgb}"
        elif t == "theme":
            return f"theme:{color.theme},tint:{color.tint}"
        elif t == "indexed":
            return f"indexed:{color.indexed}"
        else:
            return f"type={t}"
    except Exception as e:
        return f"(err:{e})"

def fill_str(fill):
    if fill is None:
        return "None"
    try:
        ft = fill.fill_type if hasattr(fill, "fill_type") else fill.patternType
        fg = color_str(fill.fgColor) if hasattr(fill, "fgColor") else "N/A"
        bg = color_str(fill.bgColor) if hasattr(fill, "bgColor") else "N/A"
        return f"fill_type={ft}, fgColor={fg}, bgColor={bg}"
    except Exception as e:
        return f"(err:{e})"

def font_str(font):
    if font is None:
        return "None"
    try:
        parts = []
        parts.append(f"name={font.name}")
        parts.append(f"size={font.size}")
        parts.append(f"bold={font.bold}")
        parts.append(f"italic={font.italic}")
        parts.append(f"color={color_str(font.color)}")
        return ", ".join(parts)
    except Exception as e:
        return f"(err:{e})"

def analyze_sheet(ws, label=""):
    print(f"\n{'='*70}")
    print(f"SHEET: {ws.title}  [{label}]")
    print(f"{'='*70}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Min row: {ws.min_row}, Max row: {ws.max_row}")
    print(f"Min col: {ws.min_column}, Max col: {ws.max_column}")

    # --- Cell values / formulas / formatting ---
    print(f"\n--- Cells ---")
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            coord = cell.coordinate
            val_str = repr(v) if v is not None else "(empty)"
            fill = fill_str(cell.fill)
            font = font_str(cell.font)
            numfmt = cell.number_format
            print(f"  [{coord}] value={val_str} | fill=({fill}) | font=({font}) | numfmt={numfmt!r}")

    # --- Merged cells ---
    print(f"\n--- Merged Cell Ranges ---")
    if ws.merged_cells.ranges:
        for rng in ws.merged_cells.ranges:
            print(f"  {rng}")
    else:
        print("  (none)")

    # --- Column widths ---
    print(f"\n--- Column Widths ---")
    if ws.column_dimensions:
        for col_letter, cd in ws.column_dimensions.items():
            print(f"  Col {col_letter}: width={cd.width}, customWidth={cd.customWidth}, hidden={cd.hidden}")
    else:
        print("  (none)")

    # --- Row heights ---
    print(f"\n--- Row Heights ---")
    if ws.row_dimensions:
        for row_idx, rd in ws.row_dimensions.items():
            print(f"  Row {row_idx}: height={rd.height}, customHeight={rd.customHeight}, hidden={rd.hidden}")
    else:
        print("  (none)")

    # --- Data validation ---
    print(f"\n--- Data Validation ---")
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            print(f"  type={dv.type}, formula1={dv.formula1}, formula2={dv.formula2}, "
                  f"sqref={dv.sqref}, showErrorMessage={dv.showErrorMessage}, "
                  f"errorTitle={dv.errorTitle}, error={dv.error}")
    else:
        print("  (none)")

# ============================================================
# Pass 1: data_only=False  (to see formulas)
# ============================================================
print("\n" + "#"*70)
print("PASS 1: data_only=False  (formulas visible)")
print("#"*70)

wb_formula = openpyxl.load_workbook(FILE, data_only=False)
print(f"\nSheet names: {wb_formula.sheetnames}")

# Named ranges
print(f"\n--- Named Ranges (workbook level) ---")
named = list(wb_formula.defined_names.definedName) if hasattr(wb_formula.defined_names, "definedName") else []
if not named:
    # Try alternative API
    try:
        named = list(wb_formula.defined_names)
    except Exception:
        named = []
if named:
    for dn in named:
        print(f"  name={dn.name!r}, value={dn.value!r}, localSheetId={dn.localSheetId}")
else:
    print("  (none)")

for ws in wb_formula.worksheets:
    analyze_sheet(ws, label="formulas")

# ============================================================
# Pass 2: data_only=True  (to see computed values)
# ============================================================
print("\n" + "#"*70)
print("PASS 2: data_only=True  (computed values)")
print("#"*70)

wb_values = openpyxl.load_workbook(FILE, data_only=True)

for ws in wb_values.worksheets:
    print(f"\n{'='*70}")
    print(f"SHEET: {ws.title}  [values]")
    print(f"{'='*70}")
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            coord = cell.coordinate
            if v is not None:
                print(f"  [{coord}] computed_value={repr(v)}")

print("\n\nDONE.")
