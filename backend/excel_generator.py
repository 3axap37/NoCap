"""
Cap Table Excel generator.

Produces an .xlsx file that exactly matches CapTableExample.xlsx in layout,
colors, merged cells, formulas, and number formats.

Strategy: load the reference file once at import time, copy cell styles
(fill, font, alignment, number_format) from specific reference cells into
the generated workbook — this preserves theme-color references so Excel
renders identical colours.
"""

import io
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from models import GenerateExcelRequest

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REFERENCE_FILE = Path(__file__).parent / "CapTableExample.xlsx"

LEAD_INVESTOR_NAME = "한투 바른동행 셰르파 제4호 펀드"

KRW_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
PCT_FMT = "0.00%"

COL_WIDTHS = {
    "A": 5.125,
    "B": 3.5,
    "C": 38.25,
    "D": 15.75,
    "E": 14.625,
    "F": 17.125,
    "G": 4.875,
    "H": 15.625,
    "I": 4.25,
}

# ---------------------------------------------------------------------------
# Load reference file (once at module level)
# ---------------------------------------------------------------------------

_ref_wb = openpyxl.load_workbook(str(REFERENCE_FILE))
_ref_ws = _ref_wb.active


def _ref(coord: str):
    """Return a reference cell from the example workbook."""
    return _ref_ws[coord]


def _apply_style(src_coord: str, dst_cell) -> None:
    """Copy fill, font, alignment, and number_format from a reference cell."""
    src = _ref(src_coord)
    if src.fill and src.fill.fill_type not in (None, "none"):
        dst_cell.fill = copy(src.fill)
    if src.font:
        dst_cell.font = copy(src.font)
    if src.alignment:
        dst_cell.alignment = copy(src.alignment)
    if src.number_format and src.number_format != "General":
        dst_cell.number_format = src.number_format


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------


def generate_excel(req: GenerateExcelRequest) -> bytes:
    N = len(req.shareholders)

    all_investors = [
        {"name": LEAD_INVESTOR_NAME, "amount": req.leadInvestorAmount}
    ] + [{"name": inv.name, "amount": inv.amount} for inv in req.coInvestors]
    M = len(all_investors)

    # ---- Row layout (dynamic) ----
    ROW_TITLE = 2
    ROW_COMPANY = 4
    ROW_ROUND = 5
    ROW_PRICE = 6
    ROW_PRE_VAL = 7
    ROW_POST_VAL = 8
    ROW_S1_LABEL = 11
    ROW_PRE_HDR = 13
    ROW_PRE_FIRST = 14
    ROW_PRE_LAST = ROW_PRE_FIRST + N - 1
    ROW_PRE_TOTAL = ROW_PRE_LAST + 1

    ROW_S2_LABEL = ROW_PRE_TOTAL + 4          # "2)" / "투자 후"
    ROW_PARAMS_END = ROW_S2_LABEL + 2 * M - 1

    ROW_POST_HDR = ROW_PARAMS_END + 2
    ROW_POST_FIRST = ROW_POST_HDR + 1
    ROW_POST_LAST_SH = ROW_POST_FIRST + N - 1
    ROW_NEW_FIRST = ROW_POST_LAST_SH + 1
    ROW_NEW_LAST = ROW_NEW_FIRST + M - 1
    ROW_POST_TOTAL = ROW_NEW_LAST + 1

    # ---- Build workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _ref_ws.title  # keep identical sheet name (Korean)

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ------------------------------------------------------------------ #
    # Title  B2                                                           #
    # ------------------------------------------------------------------ #
    ws[f"B{ROW_TITLE}"] = "Cap Table"
    _apply_style("B2", ws[f"B{ROW_TITLE}"])

    # ------------------------------------------------------------------ #
    # Company name row (B4 + C4:F4 background)                           #
    # ------------------------------------------------------------------ #
    ws[f"B{ROW_COMPANY}"] = req.companyName
    _apply_style("B4", ws[f"B{ROW_COMPANY}"])
    for col in ("C", "D", "E", "F"):
        _apply_style(f"{col}4", ws[f"{col}{ROW_COMPANY}"])

    # ------------------------------------------------------------------ #
    # Round / Purchasing Value / Pre Value / Post Value                   #
    # ------------------------------------------------------------------ #
    ws[f"C{ROW_ROUND}"] = req.round
    _apply_style("C5", ws[f"C{ROW_ROUND}"])

    ws[f"C{ROW_PRICE}"] = "Purchasing Value"
    _apply_style("C6", ws[f"C{ROW_PRICE}"])
    ws[f"D{ROW_PRICE}"] = f"=ROUND(E{ROW_PRE_VAL}/E{ROW_PRE_TOTAL},0)"
    _apply_style("D6", ws[f"D{ROW_PRICE}"])

    ws[f"C{ROW_PRE_VAL}"] = "Pre Value"
    _apply_style("C7", ws[f"C{ROW_PRE_VAL}"])
    ws[f"D{ROW_PRE_VAL}"] = f"=D{ROW_PRICE}*E{ROW_PRE_TOTAL}"
    _apply_style("D7", ws[f"D{ROW_PRE_VAL}"])
    ws[f"E{ROW_PRE_VAL}"] = req.preMoney
    _apply_style("E7", ws[f"E{ROW_PRE_VAL}"])

    ws[f"C{ROW_POST_VAL}"] = "Post Value"
    _apply_style("C8", ws[f"C{ROW_POST_VAL}"])
    ws[f"D{ROW_POST_VAL}"] = f"=D{ROW_PRICE}*E{ROW_POST_TOTAL}"
    _apply_style("D8", ws[f"D{ROW_POST_VAL}"])

    # ------------------------------------------------------------------ #
    # Section 1: 투자 전 (Pre-investment)                                  #
    # ------------------------------------------------------------------ #
    ws[f"B{ROW_S1_LABEL}"] = "1)"
    _apply_style("B11", ws[f"B{ROW_S1_LABEL}"])
    ws[f"C{ROW_S1_LABEL}"] = _ref("C11").value  # "투자 전"
    _apply_style("C11", ws[f"C{ROW_S1_LABEL}"])

    # Pre-investment header row
    for col, ref_coord in (("C", "C13"), ("D", "D13"), ("E", "E13"), ("F", "F13")):
        ws[f"{col}{ROW_PRE_HDR}"] = _ref(ref_coord).value
        _apply_style(ref_coord, ws[f"{col}{ROW_PRE_HDR}"])

    # Pre-investment shareholders
    for i, sh in enumerate(req.shareholders):
        r = ROW_PRE_FIRST + i
        ws[f"C{r}"] = sh.name
        _apply_style("C14", ws[f"C{r}"])
        ws[f"D{r}"] = sh.shareType
        _apply_style("D14", ws[f"D{r}"])
        ws[f"E{r}"] = sh.shareCount
        _apply_style("E14", ws[f"E{r}"])
        ws[f"F{r}"] = f"=E{r}/$E${ROW_PRE_TOTAL}"
        _apply_style("F14", ws[f"F{r}"])

    # Pre-investment total
    ws.merge_cells(f"C{ROW_PRE_TOTAL}:D{ROW_PRE_TOTAL}")
    ws[f"C{ROW_PRE_TOTAL}"] = _ref("C20").value  # "합계"
    _apply_style("C20", ws[f"C{ROW_PRE_TOTAL}"])
    ws[f"E{ROW_PRE_TOTAL}"] = f"=SUM(E{ROW_PRE_FIRST}:E{ROW_PRE_LAST})"
    _apply_style("E20", ws[f"E{ROW_PRE_TOTAL}"])
    ws[f"F{ROW_PRE_TOTAL}"] = f"=SUM(F{ROW_PRE_FIRST}:F{ROW_PRE_LAST})"
    _apply_style("F20", ws[f"F{ROW_PRE_TOTAL}"])

    # ------------------------------------------------------------------ #
    # Section 2: 투자 후 (Investment parameters)                           #
    # ------------------------------------------------------------------ #
    ws[f"B{ROW_S2_LABEL}"] = "2)"
    _apply_style("B24", ws[f"B{ROW_S2_LABEL}"])
    ws[f"C{ROW_S2_LABEL}"] = _ref("C24").value  # "투자 후"
    _apply_style("C24", ws[f"C{ROW_S2_LABEL}"])

    investor_param_rows: list[tuple[int, int]] = []  # (amount_row, shares_row)

    for i, investor in enumerate(all_investors):
        amount_row = ROW_S2_LABEL + 2 * i
        shares_row = ROW_S2_LABEL + 2 * i + 1
        investor_param_rows.append((amount_row, shares_row))

        ws[f"E{amount_row}"] = _ref("E24").value  # "투자금액"
        _apply_style("E24", ws[f"E{amount_row}"])
        ws[f"F{amount_row}"] = investor["amount"]
        _apply_style("F24", ws[f"F{amount_row}"])
        ws.merge_cells(f"G{amount_row}:H{amount_row}")
        _apply_style("G24", ws[f"G{amount_row}"])

        ws[f"E{shares_row}"] = _ref("E25").value  # "신주주식수"
        _apply_style("E25", ws[f"E{shares_row}"])
        ws[f"F{shares_row}"] = f"=ROUNDUP(F{amount_row}/D{ROW_PRICE},0)"
        _apply_style("F25", ws[f"F{shares_row}"])
        ws.merge_cells(f"G{shares_row}:H{shares_row}")
        _apply_style("G25", ws[f"G{shares_row}"])

    # ------------------------------------------------------------------ #
    # Post-investment header                                               #
    # ------------------------------------------------------------------ #
    for col, ref_coord in (("C", "C27"), ("D", "D27"), ("E", "E27"), ("F", "F27")):
        ws[f"{col}{ROW_POST_HDR}"] = _ref(ref_coord).value
        _apply_style(ref_coord, ws[f"{col}{ROW_POST_HDR}"])

    # ------------------------------------------------------------------ #
    # Post-investment existing shareholders                                #
    # ------------------------------------------------------------------ #
    for i, sh in enumerate(req.shareholders):
        r = ROW_POST_FIRST + i
        ws[f"C{r}"] = sh.name
        _apply_style("C28", ws[f"C{r}"])
        ws[f"D{r}"] = sh.shareType
        _apply_style("D28", ws[f"D{r}"])
        ws[f"E{r}"] = sh.shareCount
        _apply_style("E28", ws[f"E{r}"])
        ws[f"F{r}"] = f"=E{r}/$E${ROW_POST_TOTAL}"
        _apply_style("F28", ws[f"F{r}"])

    # ------------------------------------------------------------------ #
    # New RCPS investor rows                                               #
    # ------------------------------------------------------------------ #
    for i, (investor, (amount_row, shares_row)) in enumerate(
        zip(all_investors, investor_param_rows)
    ):
        new_row = ROW_NEW_FIRST + i
        ws[f"C{new_row}"] = investor["name"]
        _apply_style("C34", ws[f"C{new_row}"])
        ws[f"D{new_row}"] = "RCPS"
        _apply_style("D34", ws[f"D{new_row}"])
        ws[f"E{new_row}"] = f"=F{shares_row}"
        _apply_style("E34", ws[f"E{new_row}"])
        ws[f"F{new_row}"] = f"=E{new_row}/$E${ROW_POST_TOTAL}"
        _apply_style("F34", ws[f"F{new_row}"])
        ws.merge_cells(f"G{new_row}:H{new_row}")
        ws[f"G{new_row}"] = f"=E{new_row}*D{ROW_PRICE}"
        _apply_style("G34", ws[f"G{new_row}"])

    # ------------------------------------------------------------------ #
    # Post-investment total                                                #
    # ------------------------------------------------------------------ #
    ws.merge_cells(f"C{ROW_POST_TOTAL}:D{ROW_POST_TOTAL}")
    ws[f"C{ROW_POST_TOTAL}"] = _ref("C35").value  # "합계"
    _apply_style("C35", ws[f"C{ROW_POST_TOTAL}"])
    ws[f"E{ROW_POST_TOTAL}"] = f"=SUM(E{ROW_POST_FIRST}:E{ROW_NEW_LAST})"
    _apply_style("E35", ws[f"E{ROW_POST_TOTAL}"])
    ws[f"F{ROW_POST_TOTAL}"] = f"=SUM(F{ROW_POST_FIRST}:F{ROW_NEW_LAST})"
    _apply_style("F35", ws[f"F{ROW_POST_TOTAL}"])

    # ------------------------------------------------------------------ #
    # Serialize to bytes                                                   #
    # ------------------------------------------------------------------ #
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
